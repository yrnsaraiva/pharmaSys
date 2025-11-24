from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models import Sum, Count, Q, F
from django.contrib.auth.decorators import login_required
from vendas.models import Venda
from productos.models import Produto, Lote
from core.decorators import  vendedor_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@login_required
@vendedor_required
def dashboard(request):
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    trinta_dias_atras = hoje - timedelta(days=30)

    # Vendas do dia
    vendas_hoje = Venda.objects.filter(data_venda__date=hoje)
    total_vendas_hoje = vendas_hoje.aggregate(total=Sum('total'))['total'] or 0

    # Produtos com validade próxima (próximos 90 dias)
    validade_proxima = Lote.objects.filter(
        data_validade__range=[hoje, hoje + timedelta(days=90)]
    ).count()

    # Lista detalhada de produtos com validade próxima - COM PAGINAÇÃO
    produtos_validade_list = Lote.objects.filter(
        data_validade__range=[hoje, hoje + timedelta(days=90)]
    ).select_related('produto').order_by('data_validade')

    # Paginação - 15 itens por página
    paginator = Paginator(produtos_validade_list, 10)
    page = request.GET.get('page_validade', 1)

    try:
        produtos_validade_proxima = paginator.page(page)
    except PageNotAnInteger:
        produtos_validade_proxima = paginator.page(1)
    except EmptyPage:
        produtos_validade_proxima = paginator.page(paginator.num_pages)

    # Calcular dias para vencer para cada lote
    for lote in produtos_validade_proxima:
        lote.dias_para_vencer = (lote.data_validade - hoje).days
        lote.dias_absolutos = abs(lote.dias_para_vencer)

    # Resto do seu código...
    # Receita mensal
    receita_mensal = Venda.objects.filter(
        data_venda__date__gte=inicio_mes
    ).aggregate(total=Sum('total'))['total'] or 0

    # Vendas dos últimos 30 dias para o gráfico
    vendas_ultimos_30_dias = []
    categorias_30_dias = []

    for i in range(31):
        data = trinta_dias_atras + timedelta(days=i)
        total_dia = Venda.objects.filter(
            data_venda__date=data
        ).aggregate(total=Sum('total'))['total'] or 0

        vendas_ultimos_30_dias.append(float(total_dia))
        categorias_30_dias.append(data.strftime('%d/%m'))

    receita_ultimos_30_dias = sum(vendas_ultimos_30_dias)

    # Produtos mais vendidos (últimos 30 dias)
    produtos_mais_vendidos = Produto.objects.filter(
        itemvenda__venda__data_venda__gte=trinta_dias_atras
    ).annotate(
        total_vendido=Sum('itemvenda__quantidade')
    ).order_by('-total_vendido')[:5]

    # Preparar dados para gráfico de pizza
    produtos_pizza = []
    for produto in produtos_mais_vendidos:
        produtos_pizza.append({
            'name': produto.nome,
            'y': float(produto.total_vendido or 0)
        })

    # Productos abaixo do stock minimo
    produtos_estoque_baixo = Produto.objects.annotate(
        total_estoque=Sum('lote__quantidade_disponivel')
    ).filter(
        total_estoque__lt=F('estoque_minimo')
    )

    context = {
        'total_vendas_hoje': total_vendas_hoje,
        'estoque_baixo': produtos_estoque_baixo.count(),
        'validade_proxima': validade_proxima,
        'produtos_validade_proxima': produtos_validade_proxima,
        'receita_ultimos_30_dias': receita_ultimos_30_dias,
        'vendas_ultimos_30_dias': vendas_ultimos_30_dias,
        'categorias_30_dias': categorias_30_dias,
        'produtos_pizza': produtos_pizza,
        'hoje': hoje,
    }

    return render(request, "dashbord/dashbord.html", context)


@login_required
@vendedor_required
def exportar_validade_proxima_excel(request):
    hoje = timezone.now().date()

    # Buscar todos os produtos com validade próxima
    produtos_validade_proxima = Lote.objects.filter(
        data_validade__range=[hoje, hoje + timedelta(days=90)]
    ).select_related('produto', 'produto__categoria').order_by('data_validade')

    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos Validade Próxima"

    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Cabeçalhos ajustados para seus campos
    headers = [
        'Produto',
        'Categoria',
        'Código Barras',
        'Lote',
        'Quantidade',
        'Data Validade',
        'Dias Restantes',
        'Status',
        'Preço Venda'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Preencher dados
    for row_num, lote in enumerate(produtos_validade_proxima, 2):
        dias_para_vencer = (lote.data_validade - hoje).days

        if dias_para_vencer <= 0:
            status = "Vencido"
            dias_texto = f"Vencido há {abs(dias_para_vencer)} dias"
        elif dias_para_vencer <= 7:
            status = "Crítico"
            dias_texto = f"{dias_para_vencer} dias"
        elif dias_para_vencer <= 30:
            status = "Alerta"
            dias_texto = f"{dias_para_vencer} dias"
        else:
            status = "Atenção"
            dias_texto = f"{dias_para_vencer} dias"

        # Preencher dados com os campos corretos do seu modelo
        ws.cell(row=row_num, column=1, value=lote.produto.nome)
        ws.cell(row=row_num, column=2, value=lote.produto.categoria.nome if lote.produto.categoria else '')
        ws.cell(row=row_num, column=3, value=lote.produto.codigo_barras or '')
        ws.cell(row=row_num, column=4, value=lote.numero_lote)
        ws.cell(row=row_num, column=5, value=lote.quantidade_disponivel)
        ws.cell(row=row_num, column=6, value=lote.data_validade.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=7, value=dias_texto)
        ws.cell(row=row_num, column=8, value=status)
        ws.cell(row=row_num, column=9, value=float(lote.produto.preco_venda))

    # Ajustar largura das colunas
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                column_letter = get_column_letter(cell.column)
                current_width = column_widths.get(column_letter, 0)
                new_width = max(current_width, len(str(cell.value)) + 2)
                column_widths[column_letter] = new_width

    for column_letter, width in column_widths.items():
        ws.column_dimensions[column_letter].width = min(width, 50)

    # Criar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="produtos_validade_proxima.xlsx"'

    wb.save(response)

    return response
