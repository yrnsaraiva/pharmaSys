from django.db.models import Sum, Min, Q
from core.decorators import gerente_required, vendedor_required, admin_required
from .models import Produto, Categoria, Fornecedor, Lote
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
import datetime


# Utilitários
def get_paginated_data(queryset, request, items_per_page=10):
    """Helper para paginação com range customizado"""
    paginator = Paginator(queryset, items_per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    current_page = page_obj.number
    total_pages = paginator.num_pages
    start_page = max(current_page - 2, 1)
    end_page = min(current_page + 2, total_pages)

    return {
        "page_obj": page_obj,
        "custom_range": range(start_page, end_page + 1)
    }


def safe_decimal(value, default=0):
    """Converte seguro para Decimal"""
    if not value:
        return Decimal(default)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def safe_int(value, default=0):
    """Converte seguro para inteiro"""
    if not value:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def parse_date(date_string, default=None):
    """Converte string para date object de forma segura"""
    if not date_string:
        return default

    if isinstance(date_string, datetime.date):
        return date_string

    try:
        # Tenta converter de string no formato YYYY-MM-DD
        return datetime.datetime.strptime(date_string, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return default


# Categorias (mantidas iguais)
@login_required
@gerente_required
def categorias_list(request):
    categorias = Categoria.objects.all().order_by('nome')

    search = request.GET.get('search', '')
    if search:
        categorias = categorias.filter(nome__icontains=search)

    pagination_data = get_paginated_data(categorias, request, 5)

    return render(request, "productos/categorias.html", {
        "categorias": pagination_data["page_obj"],
        **pagination_data,
        "search": search,
    })


@login_required
@gerente_required
def criar_categoria(request):
    if request.method == "POST":
        try:
            nome = request.POST.get("nome", "").strip()
            tipo = request.POST.get("tipo")
            descricao = request.POST.get("descricao", "")

            if not nome:
                messages.error(request, "Nome da categoria é obrigatório")
                return render(request, "productos/nova_categoria.html")

            if Categoria.objects.filter(nome__iexact=nome).exists():
                messages.error(request, "Já existe uma categoria com este nome")
                return render(request, "productos/nova_categoria.html")

            categoria = Categoria.objects.create(
                nome=nome,
                tipo=tipo,
                descricao=descricao
            )

            messages.success(request, "✅ Categoria criada com sucesso!")
            return redirect('categorias_list')

        except Exception as e:
            messages.error(request, f"Erro ao criar categoria: {str(e)}")

    return render(request, "productos/nova_categoria.html")


@login_required
@gerente_required
def remover_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, pk=categoria_id)

    if Produto.objects.filter(categoria=categoria).exists():
        messages.error(request, "Não é possível excluir esta categoria pois existem produtos vinculados a ela.")
        return redirect("categorias_list")

    categoria.delete()
    messages.success(request, "✅ Categoria excluída com sucesso!")
    return redirect("categorias_list")


# Produtos (mantidas iguais)
@login_required
@vendedor_required
def productos_list(request):
    search = request.GET.get("search", "")
    categoria = request.GET.get("categoria", "Todas")
    status = request.GET.get("status", "Todos")

    productos = Produto.objects.all().select_related('categoria')

    if search:
        productos = productos.filter(
            Q(nome__icontains=search) |
            Q(codigo_barras__icontains=search) |
            Q(principio_ativo__icontains=search)
        )

    if categoria != "Todas":
        productos = productos.filter(categoria__nome=categoria)

    if status != "Todos":
        if status == "esgotado":
            productos = [p for p in productos if p.estoque_total == 0]
        elif status == "baixo":
            productos = [p for p in productos if p.status_estoque == "baixo"]
        elif status == "ok":
            productos = [p for p in productos if p.status_estoque == "ok"]

    if status != "Todos":
        paginator = Paginator(productos, 10)
    else:
        paginator = Paginator(productos, 10)

    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    custom_range = range(max(1, page_obj.number - 2), min(page_obj.number + 3, paginator.num_pages + 1))

    context = {
        "productos": page_obj,
        "page_obj": page_obj,
        "custom_range": custom_range,
        "search": search,
        "categoria": categoria,
        "status": status,
        "categorias": Categoria.objects.all()
    }
    return render(request, "productos/productos.html", context)


@login_required
@gerente_required
def cadastrar_producto(request):
    categorias = Categoria.objects.all()
    fornecedores = Fornecedor.objects.all()

    if request.method == "POST":
        try:
            nome = request.POST.get("nome", "").strip()
            categoria_id = request.POST.get("categoria")
            fornecedor_id = request.POST.get("fornecedor")
            codigo_barras = request.POST.get("codigo_barras", "").strip() or None

            preco_venda = safe_decimal(request.POST.get("preco_venda"))
            preco_compra = safe_decimal(request.POST.get("preco_compra"))
            preco_carteira = safe_decimal(request.POST.get("preco_carteira")) or None

            estoque_minimo = safe_int(request.POST.get("estoque_minimo"), 10)
            carteiras_por_caixa = safe_int(request.POST.get("carteiras_por_caixa"), 1)

            forma_farmaceutica = request.POST.get("forma_farmaceutica") or None
            dosagem = request.POST.get("dosagem", "").strip() or None
            nivel_prescricao = request.POST.get("nivel_prescricao") or None
            principio_ativo = request.POST.get("principio_ativo", "").strip() or None
            controlado = request.POST.get("controlado") == "on"

            if not nome:
                messages.error(request, "Nome do produto é obrigatório")
                return render(request, "productos/novo_produto.html", {
                    "categorias": categorias,
                    "fornecedores": fornecedores
                })

            if preco_venda <= 0:
                messages.error(request, "Preço de venda deve ser maior que zero")
                return render(request, "productos/novo_produto.html", {
                    "categorias": categorias,
                    "fornecedores": fornecedores
                })

            if preco_venda < preco_compra:
                messages.error(request, "Preço de venda não pode ser menor que preço de compra")
                return render(request, "productos/novo_produto.html", {
                    "categorias": categorias,
                    "fornecedores": fornecedores
                })

            categoria = get_object_or_404(Categoria, id=categoria_id) if categoria_id else None
            fornecedor = get_object_or_404(Fornecedor, id=fornecedor_id) if fornecedor_id else None

            produto = Produto(
                nome=nome,
                categoria=categoria,
                fornecedor=fornecedor,
                codigo_barras=codigo_barras,
                preco_venda=preco_venda,
                preco_compra=preco_compra,
                preco_carteira=preco_carteira,
                estoque_minimo=estoque_minimo,
                carteiras_por_caixa=carteiras_por_caixa,
                forma_farmaceutica=forma_farmaceutica,
                dosagem=dosagem,
                nivel_prescricao=nivel_prescricao,
                principio_ativo=principio_ativo,
                controlado=controlado,
            )

            produto.save()
            messages.success(request, "✅ Produto cadastrado com sucesso!")
            return redirect("productos_list")

        except ValidationError as e:
            messages.error(request, f"Erro de validação: {e}")
        except Exception as e:
            messages.error(request, f"Erro ao cadastrar produto: {str(e)}")

    return render(request, "productos/novo_produto.html", {
        "categorias": categorias,
        "fornecedores": fornecedores,
        "producto": {}
    })


@login_required
@admin_required
def remover_producto(request, producto_id):
    producto = get_object_or_404(Produto, pk=producto_id)

    if Lote.objects.filter(produto=producto).exists():
        messages.error(request, "Não é possível excluir este produto pois existem lotes vinculados a ele.")
        return redirect("productos_list")

    producto.delete()
    messages.success(request, "✅ Produto excluído com sucesso!")
    return redirect("productos_list")


@login_required
@admin_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(Produto, pk=producto_id)
    categorias = Categoria.objects.all()
    fornecedores = Fornecedor.objects.all()

    if request.method == "POST":
        try:
            producto.nome = request.POST.get("nome", "").strip()
            categoria_id = request.POST.get("categoria")
            fornecedor_id = request.POST.get("fornecedor")
            producto.codigo_barras = request.POST.get("codigo_barras", "").strip() or None

            producto.preco_venda = safe_decimal(request.POST.get("preco_venda"))
            producto.preco_compra = safe_decimal(request.POST.get("preco_compra"))
            preco_carteira = request.POST.get("preco_carteira")
            producto.preco_carteira = safe_decimal(preco_carteira) if preco_carteira else None

            producto.estoque_minimo = safe_int(request.POST.get("estoque_minimo"), 10)
            producto.carteiras_por_caixa = safe_int(request.POST.get("carteiras_por_caixa"), 1)

            producto.forma_farmaceutica = request.POST.get("forma_farmaceutica") or None
            producto.dosagem = request.POST.get("dosagem", "").strip() or None
            producto.nivel_prescricao = request.POST.get("nivel_prescricao") or None
            producto.principio_ativo = request.POST.get("principio_ativo", "").strip() or None
            producto.controlado = request.POST.get("controlado") == "on"

            producto.categoria = get_object_or_404(Categoria, id=categoria_id) if categoria_id else None
            producto.fornecedor = get_object_or_404(Fornecedor, id=fornecedor_id) if fornecedor_id else None

            producto.save()
            messages.success(request, "✅ Produto atualizado com sucesso!")
            return redirect('productos_list')

        except ValidationError as e:
            messages.error(request, f"Erro de validação: {e}")
        except Exception as e:
            messages.error(request, f"Erro ao atualizar produto: {str(e)}")

    context = {
        'producto': producto,
        "categorias": categorias,
        "fornecedores": fornecedores,
    }
    return render(request, 'productos/novo_produto.html', context)


# Lotes - CORRIGIDAS AS DATAS
@login_required
@gerente_required
def listar_lotes(request):
    search = request.GET.get("search", "")
    lotes = Lote.objects.select_related("produto").all().order_by('-data_criacao')

    if search:
        lotes = lotes.filter(
            Q(produto__nome__icontains=search) |
            Q(numero_lote__icontains=search)
        )

    pagination_data = get_paginated_data(lotes, request, 10)

    context = {
        "page_obj": pagination_data["page_obj"],
        **pagination_data,
        "search": search,
    }
    return render(request, "productos/lotes.html", context)


@login_required
@gerente_required
def criar_lote(request):
    if request.method == "POST":
        try:
            produto_id = request.POST.get("produto")
            produto = get_object_or_404(Produto, id=produto_id)

            nr_caixas = safe_int(request.POST.get("nr_caixas"))
            nr_carteiras = safe_int(request.POST.get("nr_carteiras"))

            data_validade = parse_date(request.POST.get("data_validade"))
            data_fabricacao = parse_date(request.POST.get("data_fabricacao")) \
                if request.POST.get("data_fabricacao") else None

            # ============================
            #   VALIDAÇÕES SIMPLIFICADAS
            # ============================
            errors = []

            if not data_validade:
                errors.append("Data de validade é obrigatória")
            elif data_validade < timezone.now().date():
                errors.append("Data de validade não pode ser no passado")

            if data_fabricacao and data_fabricacao >= data_validade:
                errors.append("Data de fabricação deve ser antes da validade")

            if nr_caixas < 0 or nr_carteiras < 0:
                errors.append("Valores não podem ser negativos")

            if nr_caixas == 0 and nr_carteiras == 0:
                errors.append("Informe pelo menos 1 caixa ou 1 carteira")

            if errors:
                for e in errors:
                    messages.error(request, e)
                return redirect("criar_lote")

            # ============================
            #     CRIA LOTE (AUTO-LOTE)
            # ============================
            lote = Lote(
                produto=produto,
                nr_caixas=nr_caixas,
                nr_carteiras=nr_carteiras,
                data_validade=data_validade,
                data_fabricacao=data_fabricacao
            )

            lote.save()  # <-- número do lote é gerado automaticamente aqui

            messages.success(request,
                             f"Lote {lote.numero_lote} criado com sucesso! "
                             f"Total: {lote.quantidade_disponivel} unidades")

            return redirect("listar_lotes")

        except Exception as e:
            messages.error(request, f"Erro ao criar lote: {str(e)}")

    context = {
        "produtos": Produto.objects.all(),
        "today": timezone.now().date()
    }
    return render(request, "productos/novo_lote.html", context)


@login_required
@gerente_required
def remover_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)

    if lote.quantidade_disponivel > 0:
        messages.warning(request,
                         f"Este lote ainda tem {lote.quantidade_disponivel} unidades em estoque. "
                         f"Tem certeza que deseja excluir?"
                         )
        return redirect("listar_lotes")

    lote.delete()
    messages.success(request, "✅ Lote excluído com sucesso!")
    return redirect("listar_lotes")

@login_required
@gerente_required
def editar_lote(request, pk):
    lote = get_object_or_404(Lote, id=pk)

    if request.method == "POST":
        try:
            lote.nr_caixas = safe_int(request.POST.get("nr_caixas"))
            lote.nr_carteiras = safe_int(request.POST.get("nr_carteiras"))

            lote.data_validade = parse_date(request.POST.get("data_validade"))
            lote.data_fabricacao = parse_date(request.POST.get("data_fabricacao")) \
                if request.POST.get("data_fabricacao") else None

            # ============================
            #       VALIDAÇÕES
            # ============================
            if not lote.data_validade:
                messages.error(request, "Data de validade é obrigatória")
                return redirect("editar_lote", pk=pk)

            if lote.data_validade < timezone.now().date():
                messages.error(request, "Data de validade não pode ser no passado")
                return redirect("editar_lote", pk=pk)

            if lote.data_fabricacao and lote.data_fabricacao >= lote.data_validade:
                messages.error(request, "Data de fabricação deve ser antes da validade")
                return redirect("editar_lote", pk=pk)

            if lote.nr_caixas < 0 or lote.nr_carteiras < 0:
                messages.error(request, "Valores não podem ser negativos")
                return redirect("editar_lote", pk=pk)

            lote.save()

            messages.success(request, "Lote atualizado com sucesso!")
            return redirect("listar_lotes")

        except Exception as e:
            messages.error(request, f"Erro ao atualizar lote: {str(e)}")

    context = {
        "lote": lote,
        "produtos": Produto.objects.all(),
        "editar": True,
        "today": timezone.now().date()
    }
    return render(request, "productos/novo_lote.html", context)




@login_required
@gerente_required
def exportar_produtos_excel(request):
    """Exporta produtos para Excel incluindo nr_caixas e nr_carteiras"""
    produtos = Produto.objects.all().order_by('categoria__nome', 'nome')

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Estoque"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=16)
    alert_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    warning_fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center')

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = "RELATÓRIO DE ESTOQUE - BALANÇO DE PRODUTOS"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Emitido em: {timezone.now().strftime('%d/%m/%Y às %H:%M')}"
    ws['A2'].alignment = center_align

    ws.append([])  # linha em branco

    # Cabeçalho da tabela com nr_caixas e nr_carteiras
    headers = [
        'Produto', 'Categoria', 'Código Barras', 'Lotes Ativos',
        'Nr Caixas', 'Nr Carteiras', 'Estoque Atual', 'Estoque Mínimo',
        'Status', 'Preço Venda'
    ]
    ws.append(headers)

    # Formata cabeçalho
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Dados dos produtos
    row_num = 5
    for produto in produtos:
        estoque_total = produto.estoque_total
        num_lotes = produto.lotes_ativos
        status_estoque = produto.status_estoque

        # Status e cor
        if status_estoque == "esgotado":
            status = "SEM ESTOQUE"
            status_fill = alert_fill
        elif status_estoque == "baixo":
            status = "ESTOQUE BAIXO"
            status_fill = warning_fill
        else:
            status = "ESTOQUE OK"
            status_fill = ok_fill

        nr_caixas = produto.lotes.filter(ativo=True).aggregate(total_caixas=Sum('nr_caixas'))['total_caixas'] or 0
        nr_carteiras = produto.lotes.filter(ativo=True).aggregate(total_carteiras=Sum('nr_carteiras'))['total_carteiras'] or 0
        
        row = [
            produto.nome,
            produto.categoria.nome if produto.categoria else 'N/A',
            produto.codigo_barras or 'N/A',
            num_lotes,
            nr_caixas,
            nr_carteiras,
            estoque_total,
            produto.estoque_minimo,
            status,
            float(produto.preco_venda) if produto.preco_venda else 0.0
        ]

        ws.append(row)

        # Formata a linha
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

            if col in [4,5,6,7,8]:  # Lotes, caixas, carteiras, estoque, mínimo
                cell.alignment = center_align
                if col in [7,8]:
                    cell.number_format = '#,##0'
            elif col == 10:  # Preço
                cell.number_format = '"MT" #,##0.00'
                cell.alignment = center_align

            if col == 9:  # Status
                cell.fill = status_fill
                cell.alignment = center_align
                cell.font = Font(bold=True)

        row_num += 1

    # Ajusta largura das colunas
    column_widths = {
        'A': 40, 'B': 20, 'C': 15, 'D': 12, 'E': 10, 'F': 12,
        'G': 15, 'H': 15, 'I': 15, 'J': 15
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Resumo do estoque
    summary_start_row = row_num + 2
    ws.merge_cells(f'A{summary_start_row}:J{summary_start_row}')
    ws[f'A{summary_start_row}'] = "RESUMO DO ESTOQUE"
    ws[f'A{summary_start_row}'].font = Font(bold=True, size=12)
    ws[f'A{summary_start_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws[f'A{summary_start_row}'].alignment = center_align

    total_sem_estoque = sum(1 for p in produtos if p.status_estoque == "esgotado")
    total_baixo_estoque = sum(1 for p in produtos if p.status_estoque == "baixo")
    total_ok_estoque = sum(1 for p in produtos if p.status_estoque == "ok")
    total_geral_estoque = sum(p.estoque_total for p in produtos)

    resumo_data = [
        ['Total de Produtos', len(produtos)],
        ['Produtos sem Estoque', total_sem_estoque],
        ['Produtos com Estoque Baixo', total_baixo_estoque],
        ['Produtos com Estoque OK', total_ok_estoque],
        ['Estoque Total Geral', f"{total_geral_estoque} unidades"],
        ['Rendimento Potencial Total', sum(p.valor_investido_total for p in produtos)],
        ['Investimento', sum(p.rendimento_total for p in produtos)],
    ]

    for i, (descricao, valor) in enumerate(resumo_data, start=1):
        row_idx = summary_start_row + i
        cell_desc = ws.cell(row=row_idx, column=1, value=descricao)
        cell_desc.font = Font(bold=True)
        cell_desc.border = thin_border
        cell_val = ws.cell(row=row_idx, column=2, value=valor)
        cell_val.border = thin_border
        if i < len(resumo_data):
            cell_val.number_format = '#,##0'

    # Prepara resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_estoque.xlsx"'
    wb.save(response)
    return response
